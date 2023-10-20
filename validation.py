import holidays
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def remove_weekends_and_holidays(df):
    
    df['Data2'] = pd.to_datetime(df['Data'], format='%d/%m/%Y')
    
    df['Remover'] = df['Data2'].apply(validation_weekend_and_holidays)

    # Remova os registros que atendem à condição (Remover = True)
    df = df[~df['Remover']]

    # Remova a coluna 'Remover' se não for mais necessária
    df = df.drop('Remover', axis=1)
    df = df.drop('Data2', axis=1)
    
def sum_by_hours_and_minutes(row):
    horas = 0
    minutos = 0
    for value in row:
        if ':' in value:
            parts = value.split(':')
            if len(parts) == 2:
                horas += int(parts[0])
                minutos += int(parts[1])
    # Realizar ajuste se os minutos ultrapassarem 60
    horas += minutos // 60
    minutos %= 60
    return f"{horas:02}:{minutos:02}"

def validation_weekend_and_holidays(date):
    
    dia_da_semana = date.weekday()

    pais = "BR"

    # Verifique se a data é um feriado
    if date in holidays.CountryHoliday(pais, observed=True) or (dia_da_semana == 5 or dia_da_semana == 6):
        return True
    else:
        return False

def adjust_lunch_time(df, hora_almoco):
    # Separe 'Hour' e 'Minute' da coluna 'Total' e converta em inteiros
    df[['Hour', 'Minute']] = df['Total'].str.split(":", n=1, expand=True).astype(int)

    # Faça o ajuste necessário com a hora do almoço
    df['Hour'] = df['Hour'] - hora_almoco

    # Converta 'Hour' e 'Minute' de volta em uma string no formato HH:MM e atribua a 'Total'
    df['Total'] = df['Hour'].astype(str) + ":" + df['Minute'].astype(str).str.zfill(2)

def generate_excel(df):
    fileExcel = 'Output.xlsx'
    df = df.fillna('0:00')
    
    df['Total por Linha'] = df.apply(sum_by_hours_and_minutes, axis=1)

    # Aplicar a função nas colunas do DataFrame
    df.loc['Total por Coluna'] = df.apply(sum_by_hours_and_minutes, axis=0)
    
    # Salve o DataFrame em um novo arquivo Excel
    df.to_excel(fileExcel, sheet_name='Data', index=True)

    # Carregue o arquivo Excel com a nova planilha
    wb = load_workbook(fileExcel)

    # Obtenha a planilha 'Data'
    ws = wb['Data']

    # Defina uma cor de preenchimento vermelho
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Itere pelas células na planilha e verifique se o valor é "9:00"
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row - 1, max_col=ws.max_column - 1):
        for cell in row:
            if pd.notna(cell.value) and ':' in cell.value:
                    parts = cell.value.split(':')
                    if len(parts) > 1:
                        horas = int(parts[0])
                        minutos = int(parts[1])
                        if horas > 8 or (horas == 8 and minutos > 0):
                            cell.fill = red_fill
                    
    wb.save(fileExcel)
    
#TODO verificar a regra correta para a conversão
def converter_valor(valor):
    parts = valor.split(':')
    valor = float(parts[0])
    
    if valor < 3:
        return 0
    elif 3 <= valor <= 6:
        return 0.5
    else:
        return 1    

def Update_Format_Hours_To_Points():
    fileExcel = 'Output.xlsx'
    workbook = load_workbook(fileExcel)
    aba = workbook['Data']
    
    dados = []
    header = []
    for i, row in enumerate(aba.iter_rows(min_row=1, max_row=aba.max_row - 1, min_col=1, max_col=aba.max_column - 1)):
        row_data = [cell.value for cell in row]
        if i != 0: 
            dados.append(row_data)
        else:
            header = row_data

    # Converter os dados em um DataFrame
    df = pd.DataFrame(dados, columns=header)

    nova_aba = workbook.create_sheet('NomeDaNovaAba')
    
    # Aplicar a função de conversão às colunas de datas
    df.iloc[:, 1:] = df.iloc[:, 1:].applymap(converter_valor)

    # # Soma por linha nas colunas de datas
    df['Total por Linha'] = df.iloc[:, 1:].sum(axis=1)
    
    # Escrever os dados do DataFrame na nova aba
    for row in dataframe_to_rows(df, index=False, header=True):
        nova_aba.append(row)

    # Excluir a aba original
    # workbook.remove(workbook['NomeDaSuaAba'])

    # Salvar a planilha com a nova aba e os dados
    workbook.save(fileExcel)

    