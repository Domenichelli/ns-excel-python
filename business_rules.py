import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def soma_horas_minutos2(row):
    horas = 0
    minutos = 0
    for value in row:
        if ':' in value:
            parts = value.split(':')
            if len(parts) == 2:
                horas += int(parts[0])
                minutos += int(parts[1])
    # Realizar ajuste se os minutos ultrapassarem 60
    horas += minutos // 60
    minutos %= 60
    return f"{horas:02}:{minutos:02}"

def processar_arquivo_excel(arquivo_excel, colunas_desejadas):
    df = pd.read_excel(
        arquivo_excel,
        usecols=colunas_desejadas
    )

    df = df.drop_duplicates(subset=['Matricula', 'Nome', 'Data'], keep='last')

    df['indice'] = range(len(df))

    # Defina a nova coluna como o índice do DataFrame
    df.set_index('indice', inplace=True)
    
    pivot_table = pd.pivot_table(df, values=['Total'], index=['Nome'], columns=['Data'], aggfunc='sum')
    
    pivot_table.index.name = None
    pivot_table.columns.name = None
    
    pivot_table = pivot_table.fillna('0:00')
    
    pivot_table['Total por Linha'] = pivot_table.apply(soma_horas_minutos2, axis=1)

    # Aplicar a função nas colunas do DataFrame
    pivot_table.loc['Total por Coluna'] = pivot_table.apply(soma_horas_minutos2, axis=0)
    
    with pd.ExcelWriter('Output.xlsx', engine='openpyxl') as writer:
        pivot_table.to_excel(writer, sheet_name='Planilha', startrow=0, header=True, index=True)
        worksheet = writer.sheets['Planilha']
        worksheet.delete_rows(3)
        
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for row in worksheet.iter_rows(min_row=2, max_row=len(pivot_table) + 1, min_col=2, max_col=len(pivot_table.columns)):
            for cell in row:
                if pd.notna(cell.value) and ':' in cell.value:
                    parts = cell.value.split(':')
                    if len(parts) > 1:
                        horas = int(parts[0])
                        minutos = int(parts[1])
                        if horas > 5 or (horas == 5 and minutos > 17):
                            cell.fill = red_fill
        
        # Salvar o arquivo Excel
        writer._save()

    # print(pivot_table)
    return df.head(100)


# Chame a função com os parâmetros necessários
arquivo_excel = 'EMTHOS.xlsx'
colunas_desejadas = ['Matricula', 'Nome', 'Data', 'Total']

df_resultado = processar_arquivo_excel(arquivo_excel, colunas_desejadas)